"""Convert each vocabulary PDF into its own Excel and app-ready CSV file.

Usage:
  python tools/pdf_vocab_to_excel.py
  python tools/pdf_vocab_to_excel.py --input "vocab pdf" --output-dir "exports/by_topic"

The generated CSV uses the import header expected by Local English Flashcards:
word,ipa,meaningVi,definitionEn,exampleEn,exampleVi,partOfSpeech,level
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_COLUMNS = [
    "word",
    "ipa",
    "meaningVi",
    "definitionEn",
    "exampleEn",
    "exampleVi",
    "partOfSpeech",
    "level",
]

EXCEL_COLUMNS = [
    "topic",
    "sourceFile",
    "itemNumber",
    *APP_COLUMNS,
]


@dataclass
class VocabRow:
    topic: str
    source_file: str
    item_number: int
    word: str
    ipa: str
    meaning_vi: str
    definition_en: str
    example_en: str
    example_vi: str
    part_of_speech: str = "phrase"
    level: str = "IELTS"

    def excel_values(self) -> list[str | int]:
        return [
            self.topic,
            self.source_file,
            self.item_number,
            self.word,
            self.ipa,
            self.meaning_vi,
            self.definition_en,
            self.example_en,
            self.example_vi,
            self.part_of_speech,
            self.level,
        ]

    def csv_values(self) -> list[str]:
        return [
            self.word,
            self.ipa,
            self.meaning_vi,
            self.definition_en,
            self.example_en,
            self.example_vi,
            self.part_of_speech,
            self.level,
        ]


def clean_text(value: str) -> str:
    value = value.replace("\u00a0", " ").replace("’", "'")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\s+\n", "\n", value)
    value = re.sub(r"\n\s+", "\n", value)
    return value.strip()


def topic_from_filename(path: Path) -> str:
    stem = re.sub(r"\s*\(\d+\)$", "", path.stem)
    stem = re.sub(r"^\d+\.", "", stem)
    return stem.replace("_", " ").strip().title()


def natural_key(path: Path) -> list[int | str]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", path.name)]


def safe_filename(path: Path) -> str:
    stem = re.sub(r"\s*\(\d+\)$", "", path.stem)
    stem = stem.lower().replace("&", "and")
    stem = re.sub(r"[^a-z0-9]+", "_", stem)
    return re.sub(r"_+", "_", stem).strip("_")


def pdf_text(path: Path) -> str:
    doc = fitz.open(path)
    return clean_text("\n".join(page.get_text("text") for page in doc))


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def section_value(block: str, label: str, next_labels: list[str]) -> str:
    if not next_labels:
        match = re.search(rf"{re.escape(label)}\s*(.*?)\Z", block, re.IGNORECASE | re.DOTALL)
        return clean_text(match.group(1)) if match else ""

    stop = "|".join(re.escape(item) for item in next_labels)
    pattern = rf"{re.escape(label)}\s*(.*?)(?=\n(?:{stop})|\Z)"
    match = re.search(pattern, block, re.IGNORECASE | re.DOTALL)
    return clean_text(match.group(1)) if match else ""


def parse_translation(block: str, english_word: str) -> str:
    tail_match = re.search(r"Dịch đại\s*ý\s*(.*?)(?=\n(?:Nghĩa:|Ví dụ:)|\Z)", block, re.IGNORECASE | re.DOTALL)
    if not tail_match:
        return ""

    tail = clean_text(tail_match.group(1))
    # Most PDFs repeat the phrase as "word: Vietnamese meaning".
    if ":" in tail:
        left, right = tail.split(":", 1)
        if clean_text(left).lower() in clean_text(english_word).lower() or clean_text(english_word).lower() in clean_text(left).lower():
            return clean_text(right)
        return clean_text(right)
    return tail


def parse_blocks(text: str) -> list[tuple[int, str]]:
    heading = re.compile(r"(?m)^\s*(\d+)\.\s+(.+?)\s*$")
    matches = list(heading.finditer(text))
    blocks: list[tuple[int, str]] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        blocks.append((int(match.group(1)), text[start:end]))
    return blocks


def parse_pdf(path: Path) -> list[VocabRow]:
    text = pdf_text(path)
    topic = topic_from_filename(path)
    rows: list[VocabRow] = []

    for item_number, block in parse_blocks(text):
        first_line = block.splitlines()[0]
        word = clean_text(re.sub(r"^\s*\d+\.\s+", "", first_line))
        definition_en = section_value(block, "Meaning:", ["Example:", "Dịch đại ý", "Dịch đại  ý", "Nghĩa:", "Ví dụ:"])
        example_en = section_value(block, "Example:", ["Dịch đại ý", "Dịch đại  ý", "Nghĩa:", "Ví dụ:"])
        meaning_vi = parse_translation(block, word)
        example_vi = section_value(block, "Ví dụ:", [])

        rows.append(
            VocabRow(
                topic=topic,
                source_file=path.name,
                item_number=item_number,
                word=word,
                ipa="",
                meaning_vi=meaning_vi,
                definition_en=definition_en,
                example_en=example_en,
                example_vi=example_vi,
            )
        )

    return rows


def autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 55)


def write_workbook(rows: list[VocabRow], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Vocabulary"
    ws.append(EXCEL_COLUMNS)
    for row in rows:
        ws.append(row.excel_values())

    import_ws = wb.create_sheet("App Import")
    import_ws.append(APP_COLUMNS)
    for row in rows:
        import_ws.append(row.csv_values())

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_sheet(sheet)

    wb.save(output)


def write_csv(rows: list[VocabRow], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(APP_COLUMNS)
        for row in rows:
            writer.writerow(row.csv_values())


def collect_pdf_rows(input_dir: Path) -> tuple[list[tuple[Path, list[VocabRow]]], list[str]]:
    pdf_outputs: list[tuple[Path, list[VocabRow]]] = []
    warnings: list[str] = []
    seen_hashes: set[str] = set()

    for path in sorted(input_dir.glob("*.pdf"), key=natural_key):
        digest = file_hash(path)
        if digest in seen_hashes:
            warnings.append(f"Skipped duplicate PDF content: {path.name}")
            continue
        seen_hashes.add(digest)

        pdf_rows = parse_pdf(path)
        if not pdf_rows:
            warnings.append(f"No vocabulary rows found: {path.name}")
            continue
        pdf_outputs.append((path, pdf_rows))

    return pdf_outputs, warnings


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert each vocabulary PDF to a separate Excel and CSV file.")
    parser.add_argument("--input", default="vocab pdf", help="Folder containing vocabulary PDF files.")
    parser.add_argument("--output-dir", default="exports/by_topic", help="Folder for per-topic Excel and CSV files.")
    parser.add_argument("--combined", action="store_true", help="Also write combined Excel/CSV files for review only.")
    args = parser.parse_args()

    input_dir = Path(args.input)
    if not input_dir.exists():
        raise SystemExit(f"Input folder not found: {input_dir}")

    pdf_outputs, warnings = collect_pdf_rows(input_dir)
    if not pdf_outputs:
        raise SystemExit("No vocabulary rows were extracted.")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    for pdf_path, rows in pdf_outputs:
        total_rows += len(rows)
        base = safe_filename(pdf_path)
        xlsx_path = output_dir / f"{base}.xlsx"
        csv_path = output_dir / f"{base}.csv"
        write_workbook(rows, xlsx_path)
        write_csv(rows, csv_path)
        print(f"{pdf_path.name}: {len(rows)} rows")
        print(f"  Excel: {xlsx_path.resolve()}")
        print(f"  CSV:   {csv_path.resolve()}")

    if args.combined:
        combined_rows = [row for _, rows in pdf_outputs for row in rows]
        write_workbook(combined_rows, output_dir / "_combined_vocab_import.xlsx")
        write_csv(combined_rows, output_dir / "_combined_vocab_import.csv")
        print("Combined review files written with _combined_ prefix.")

    print(f"Extracted {total_rows} vocabulary rows from {len(pdf_outputs)} PDF files.")
    for warning in warnings:
        print(f"Warning: {warning}")


if __name__ == "__main__":
    main()
